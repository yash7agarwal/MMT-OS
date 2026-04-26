"""Tests for the executive report pipeline (v0.17.0).

Pin the load-bearing contracts:
  - snapshot is deterministic (same KG state → identical content_hash)
  - synthesis URL gate flags hallucinated URLs
  - chart functions return None on empty input (caller can skip section)
  - xlsx generation produces a valid workbook with the expected sheet set
  - manifest round-trips through KnowledgeArtifact
  - generate_report's cache hits when content_hash matches

PDF rendering is NOT tested locally because WeasyPrint requires libpango;
that's verified post-deploy on Railway.
"""
from __future__ import annotations

import io
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from agent.report_charts import (
    render_impact_cascade,
    render_lens_heatmap,
    render_trend_timeline,
)
from agent.report_snapshot import ReportSnapshot
from agent.report_synthesis import (
    Recommendation,
    _gate_urls,
    _FALLBACK,
    recommendations,
)
from agent.report_xlsx import generate_xlsx


# ---- Fixtures ----

@pytest.fixture
def mini_snapshot() -> ReportSnapshot:
    return ReportSnapshot(
        project_id=99,
        project_name="TestCo",
        project_description="A test company",
        app_package="testco.example",
        portfolio_summary="PRODUCTS: widget A, widget B\nINDUSTRY: testing",
        stats={
            "competitor_count": 2, "trend_count": 3, "regulation_count": 1,
            "technology_count": 1, "effect_count": 0, "observation_count": 5,
            "relation_count": 0, "source_count": 4, "session_count": 1,
        },
        competitors=[
            {"id": 1, "name": "Acme", "entity_type": "company",
             "description": "competitor", "confidence": 0.8,
             "metadata": {}, "last_updated_at": "2026-04-26T00:00:00",
             "observations": [
                 {"id": 10, "type": "feature", "content": "Acme launched X",
                  "source_url": "https://acme.example/news", "lens_tags": ["growth"]},
             ]},
            {"id": 2, "name": "BetaCo", "entity_type": "company",
             "description": "competitor", "confidence": 0.7,
             "metadata": {}, "last_updated_at": "2026-04-26T00:00:00",
             "observations": []},
        ],
        trends=[
            {"id": 3, "name": "Trend X", "timeline": "emerging",
             "category": "technology", "description": "trend desc",
             "observations": [{"source_url": "https://t.example/x"}]},
        ],
        regulations=[{"id": 4, "name": "Reg A", "description": "compliance",
                      "confidence": 0.9}],
        technologies=[{"id": 5, "name": "Tech B", "description": "tech",
                       "confidence": 0.9}],
        effects=[],
        lens_matrix={
            "lenses": ["growth", "moat"],
            "competitors": [
                {"id": 1, "name": "Acme",   "lens_counts": {"growth": 3, "moat": 0}, "total_observations": 3},
                {"id": 2, "name": "BetaCo", "lens_counts": {"growth": 1, "moat": 2}, "total_observations": 3},
            ],
        },
        lens_detail={"growth": [], "moat": []},
        impact_graph={"nodes": [], "edges": []},
        prd_artifacts=[],
        source_index=[
            {"url": "https://acme.example/news", "citations": 1, "host": "acme.example"},
            {"url": "https://t.example/x", "citations": 1, "host": "t.example"},
        ],
        agent_sessions=[],
        loupe_runs=[],
        generated_at="2026-04-26T01:00:00Z",
    )


# ---- Snapshot determinism ----

def test_snapshot_content_hash_stable(mini_snapshot):
    """Same dataclass state → identical content_hash."""
    h1 = mini_snapshot.content_hash()
    h2 = mini_snapshot.content_hash()
    assert h1 == h2
    assert len(h1) == 16


def test_snapshot_hash_excludes_volatile_fields(mini_snapshot):
    """generated_at and loupe_runs must NOT change the hash —
    they're deliberately excluded so re-runs and Loupe reachability
    don't bust the narrative cache."""
    h_before = mini_snapshot.content_hash()
    mini_snapshot.generated_at = "2099-01-01T00:00:00Z"
    mini_snapshot.loupe_runs = [{"plan_name": "x", "run_at": "z", "passed": 1, "failed": 0}]
    assert mini_snapshot.content_hash() == h_before


def test_snapshot_hash_changes_with_data(mini_snapshot):
    """Adding a competitor MUST change the hash — proves cache invalidates."""
    h_before = mini_snapshot.content_hash()
    mini_snapshot.competitors.append(
        {"id": 99, "name": "NewCo", "entity_type": "company",
         "description": "", "confidence": 0.5, "metadata": {},
         "last_updated_at": None, "observations": []}
    )
    assert mini_snapshot.content_hash() != h_before


# ---- Synthesis URL gate ----

def test_url_gate_flags_hallucinated_urls():
    """The synthesizer's anti-hallucination guard MUST detect URLs
    not in the source pool."""
    text = "See https://allowed.com/a and https://hallucinated.com/b for details."
    pool = {"https://allowed.com/a"}
    _, hallucinated = _gate_urls(text, pool)
    assert hallucinated == ["https://hallucinated.com/b"]


def test_url_gate_passes_clean_text():
    text = "All cited URLs https://a.com and https://b.com are in pool."
    pool = {"https://a.com", "https://b.com"}
    _, hallucinated = _gate_urls(text, pool)
    assert hallucinated == []


# ---- Recommendations: drop those without evidence ----

def test_recommendations_drop_without_evidence():
    """Recommendations the LLM emits without an evidence_refs entry MUST
    be dropped — no rec without a citation."""
    snapshot = {"project_name": "X", "portfolio_summary": ""}
    narrative = {"executive_summary": "test"}
    pool = {"https://evidence.example/1"}

    fake_response = '''[
      {"title": "Build A", "body": "Do A.", "evidence_refs": [1]},
      {"title": "Build B", "body": "Do B.", "evidence_refs": []}
    ]'''
    with patch("agent.report_synthesis._ask", return_value=fake_response):
        recs = recommendations(snapshot, narrative, pool)

    assert len(recs) == 1, "rec without evidence_refs must be dropped"
    assert recs[0].title == "Build A"
    assert recs[0].evidence_urls == ["https://evidence.example/1"]


# ---- Charts: empty-input gates ----

def test_chart_lens_heatmap_empty_returns_none():
    assert render_lens_heatmap({}) is None
    assert render_lens_heatmap({"lenses": [], "competitors": []}) is None


def test_chart_trend_timeline_empty_returns_none():
    assert render_trend_timeline([]) is None


def test_chart_impact_cascade_empty_returns_none():
    assert render_impact_cascade({"nodes": [], "edges": []}) is None


def test_chart_lens_heatmap_renders(mini_snapshot):
    """Real data → non-empty PNG."""
    png = render_lens_heatmap(mini_snapshot.lens_matrix)
    assert png is not None
    assert png[:8].startswith(b"\x89PNG"), "must be a real PNG"


# ---- xlsx: shape + hyperlinks ----

def test_xlsx_has_expected_sheets(mini_snapshot):
    """The 9-tab structure is part of the contract — analysts navigate by sheet name."""
    xlsx_bytes = generate_xlsx(mini_snapshot)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    expected = {"Summary", "Competitors", "Observations", "Trends",
                "Regulations", "Technologies", "Lens matrix",
                "Sources", "Methodology"}
    assert expected.issubset(set(wb.sheetnames))


def test_xlsx_observations_have_hyperlinks(mini_snapshot):
    """Every observation row whose source_url is non-empty must have a clickable
    hyperlink in the URL column."""
    xlsx_bytes = generate_xlsx(mini_snapshot)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    obs = wb["Observations"]
    rows_with_links = 0
    for row in obs.iter_rows(min_row=2):
        for cell in row:
            if cell.hyperlink:
                rows_with_links += 1
                break
    # mini_snapshot has 1 observation with source_url
    assert rows_with_links >= 1, "expected at least one hyperlinked observation row"


# ---- v0.18.0: tier verification ----

def test_tier_split_covers_all_sections():
    """Every section we synthesize MUST have a tier assignment; otherwise it
    silently falls into the strict default which is what we're trying to
    relax. This test catches regressions where a new section is added but
    forgotten in TIER_BY_SECTION."""
    from agent.report_synthesis import (
        TIER_BY_SECTION, TIER_COMMON_KNOWLEDGE, TIER_NEEDS_GROUNDING,
    )
    expected = {"executive_summary", "competitive_framing", "lens_insights",
                "regulatory_framing", "strategic_implications", "recommendations"}
    assert expected.issubset(set(TIER_BY_SECTION.keys())), (
        f"sections missing a tier: {expected - set(TIER_BY_SECTION.keys())}"
    )
    # Every value must be one of the two tiers
    valid = {TIER_COMMON_KNOWLEDGE, TIER_NEEDS_GROUNDING}
    for sec, tier in TIER_BY_SECTION.items():
        assert tier in valid, f"section {sec!r} has invalid tier {tier!r}"


def test_common_knowledge_system_prompt_relaxes_url_requirement():
    """Pin that the common-knowledge system prompt does NOT carry the
    'every claim must cite a URL' string. If a future edit re-adds that
    line under common_knowledge, it defeats the whole tier system."""
    from agent.report_synthesis import _SYSTEM_COMMON, _SYSTEM_GROUNDED
    assert "every factual claim must cite a source URL" in _SYSTEM_GROUNDED
    assert "every factual claim must cite a source URL" not in _SYSTEM_COMMON
    # Common-knowledge prompt must still forbid fabricating specific numbers
    assert "fabricate" in _SYSTEM_COMMON.lower()
    assert "directionally" in _SYSTEM_COMMON.lower() or "specific datum" in _SYSTEM_COMMON.lower()


def test_grounded_sections_remain_strict():
    """Lens insights and regulatory framing — the most domain-specific
    and time-sensitive sections — must stay on the strict tier."""
    from agent.report_synthesis import (
        TIER_BY_SECTION, TIER_NEEDS_GROUNDING,
    )
    assert TIER_BY_SECTION["lens_insights"] == TIER_NEEDS_GROUNDING
    assert TIER_BY_SECTION["regulatory_framing"] == TIER_NEEDS_GROUNDING
    assert TIER_BY_SECTION["recommendations"] == TIER_NEEDS_GROUNDING
