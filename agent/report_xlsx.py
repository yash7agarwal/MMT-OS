"""Excel (.xlsx) generator for the executive report (v0.17.0).

Companion to the PDF — same KG snapshot, but exported as a tabbed
spreadsheet for analyst deep-dive. Every observation row has a
clickable hyperlink to its source URL; the lens matrix tab uses
openpyxl conditional formatting for a printable heatmap.

Returns xlsx bytes. Single function `generate_xlsx(snapshot) -> bytes`.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agent.report_snapshot import ReportSnapshot

logger = logging.getLogger(__name__)


# Brand palette (hex without #)
ZINC_950 = "FF09090B"
ZINC_300 = "FFD4D4D8"
ZINC_100 = "FFF4F4F5"
EMERALD_500 = "FF10B981"
EMERALD_700 = "FF047857"

# Styles
HEADER_FILL = PatternFill(start_color="FF18181B", end_color="FF18181B", fill_type="solid")
HEADER_FONT = Font(name="Inter", size=10, bold=True, color="FFFFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
BODY_FONT = Font(name="Inter", size=10, color="FF18181B")
URL_FONT = Font(name="Inter", size=9, color="FF047857", underline="single")
THIN_BORDER = Border(bottom=Side(style="thin", color="FFE4E4E7"))


def _setup_sheet(ws, headers: list[str], col_widths: list[int]):
    """Common sheet styling — black header bar, freeze top row."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _link_cell(ws, row: int, col: int, url: str, display: str | None = None):
    """Write a clickable hyperlink in a cell."""
    if not url:
        ws.cell(row=row, column=col, value="").font = BODY_FONT
        return
    cell = ws.cell(row=row, column=col, value=display or url[:80])
    cell.hyperlink = url
    cell.font = URL_FONT


def generate_xlsx(snapshot: ReportSnapshot) -> bytes:
    """Build the multi-tab workbook and return bytes."""
    wb = Workbook()

    # Tab 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    rows = [
        ("Project", snapshot.project_name),
        ("Generated at", snapshot.generated_at),
        ("Snapshot hash", snapshot.content_hash()),
        ("App package / URL", snapshot.app_package or "—"),
        ("Description", snapshot.project_description or "—"),
        ("", ""),
        ("Competitor count", snapshot.stats.get("competitor_count", 0)),
        ("Trend count", snapshot.stats.get("trend_count", 0)),
        ("Regulation count", snapshot.stats.get("regulation_count", 0)),
        ("Technology count", snapshot.stats.get("technology_count", 0)),
        ("Effect count", snapshot.stats.get("effect_count", 0)),
        ("Observation count", snapshot.stats.get("observation_count", 0)),
        ("Unique source URLs", snapshot.stats.get("source_count", 0)),
        ("Agent sessions", snapshot.stats.get("session_count", 0)),
    ]
    for r_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=label).font = Font(name="Inter", size=10, bold=True, color="FF18181B")
        ws.cell(row=r_idx, column=2, value=value).font = BODY_FONT

    # Tab 2: Competitors
    ws = wb.create_sheet("Competitors")
    headers = ["Name", "Description", "Confidence", "Last updated"]
    _setup_sheet(ws, headers, [28, 60, 12, 18])
    for r_idx, c in enumerate(snapshot.competitors, start=2):
        ws.cell(row=r_idx, column=1, value=c.get("name", "")).font = BODY_FONT
        ws.cell(row=r_idx, column=2, value=(c.get("description") or "")[:300]).font = BODY_FONT
        ws.cell(row=r_idx, column=3, value=round(c.get("confidence") or 0, 2)).font = BODY_FONT
        ws.cell(row=r_idx, column=4, value=(c.get("last_updated_at") or "")[:10]).font = BODY_FONT

    # Tab 3: Observations (every finding with hyperlink) — the analyst's home page
    ws = wb.create_sheet("Observations")
    headers = ["Entity", "Type", "Lens tags", "Content", "Source URL", "Recorded"]
    _setup_sheet(ws, headers, [22, 14, 22, 80, 50, 18])
    obs_row = 2
    # Build a flat observation list from competitor + lens data
    seen = set()  # de-dup by observation id
    for entity in snapshot.competitors:
        for o in entity.get("observations", []) or []:
            oid = o.get("id")
            if oid in seen:
                continue
            seen.add(oid)
            ws.cell(row=obs_row, column=1, value=entity.get("name", "")).font = BODY_FONT
            ws.cell(row=obs_row, column=2, value=o.get("observation_type") or o.get("type") or "").font = BODY_FONT
            tags = o.get("lens_tags") or []
            if isinstance(tags, list):
                ws.cell(row=obs_row, column=3, value=", ".join(tags)).font = BODY_FONT
            ws.cell(row=obs_row, column=4, value=(o.get("content") or "")[:500]).font = BODY_FONT
            _link_cell(ws, obs_row, 5, o.get("source_url") or "")
            ws.cell(row=obs_row, column=6, value=(o.get("recorded_at") or o.get("observed_at") or "")[:16]).font = BODY_FONT
            obs_row += 1
    # Also pull from lens_detail observations (covers per-lens drilldown)
    for lens, entities in (snapshot.lens_detail or {}).items():
        for entity in entities:
            for o in entity.get("observations", []) or []:
                oid = o.get("id")
                if oid in seen:
                    continue
                seen.add(oid)
                ws.cell(row=obs_row, column=1, value=entity.get("name", "")).font = BODY_FONT
                ws.cell(row=obs_row, column=2, value=o.get("observation_type") or "").font = BODY_FONT
                ws.cell(row=obs_row, column=3, value=lens).font = BODY_FONT
                ws.cell(row=obs_row, column=4, value=(o.get("content") or "")[:500]).font = BODY_FONT
                _link_cell(ws, obs_row, 5, o.get("source_url") or "")
                ws.cell(row=obs_row, column=6, value=(o.get("recorded_at") or o.get("observed_at") or "")[:16]).font = BODY_FONT
                obs_row += 1

    # Tab 4: Trends
    ws = wb.create_sheet("Trends")
    headers = ["Name", "Timeline", "Category", "Description", "Top source"]
    _setup_sheet(ws, headers, [40, 12, 16, 70, 40])
    for r_idx, t in enumerate(snapshot.trends, start=2):
        ws.cell(row=r_idx, column=1, value=t.get("name", "")[:120]).font = BODY_FONT
        ws.cell(row=r_idx, column=2, value=t.get("timeline", "")).font = BODY_FONT
        ws.cell(row=r_idx, column=3, value=t.get("category", "")).font = BODY_FONT
        ws.cell(row=r_idx, column=4, value=(t.get("description") or "")[:300]).font = BODY_FONT
        first_obs = (t.get("observations") or [None])[0]
        url = first_obs.get("source_url", "") if first_obs else ""
        _link_cell(ws, r_idx, 5, url)

    # Tab 5: Regulations
    ws = wb.create_sheet("Regulations")
    headers = ["Name", "Description", "Confidence"]
    _setup_sheet(ws, headers, [40, 80, 12])
    for r_idx, r in enumerate(snapshot.regulations, start=2):
        ws.cell(row=r_idx, column=1, value=r.get("name", "")[:120]).font = BODY_FONT
        ws.cell(row=r_idx, column=2, value=(r.get("description") or "")[:400]).font = BODY_FONT
        ws.cell(row=r_idx, column=3, value=round(r.get("confidence") or 0, 2)).font = BODY_FONT

    # Tab 6: Technologies
    ws = wb.create_sheet("Technologies")
    headers = ["Name", "Description", "Confidence"]
    _setup_sheet(ws, headers, [40, 80, 12])
    for r_idx, t in enumerate(snapshot.technologies, start=2):
        ws.cell(row=r_idx, column=1, value=t.get("name", "")[:120]).font = BODY_FONT
        ws.cell(row=r_idx, column=2, value=(t.get("description") or "")[:400]).font = BODY_FONT
        ws.cell(row=r_idx, column=3, value=round(t.get("confidence") or 0, 2)).font = BODY_FONT

    # Tab 7: Lens matrix (heatmap conditional formatting)
    ws = wb.create_sheet("Lens matrix")
    matrix = snapshot.lens_matrix or {}
    lenses = matrix.get("lenses") or []
    competitors = matrix.get("competitors") or []
    if lenses and competitors:
        headers = ["Lens"] + [c.get("name", "")[:24] for c in competitors]
        col_widths = [22] + [16] * len(competitors)
        _setup_sheet(ws, headers, col_widths)
        for r_idx, lens in enumerate(lenses, start=2):
            ws.cell(row=r_idx, column=1, value=lens.replace("_", " ").title()).font = Font(name="Inter", size=10, bold=True)
            for c_idx, comp in enumerate(competitors, start=2):
                v = (comp.get("lens_counts") or {}).get(lens, 0)
                ws.cell(row=r_idx, column=c_idx, value=v).font = BODY_FONT
        # Heatmap rule across the data range
        last_col = get_column_letter(1 + len(competitors))
        last_row = 1 + len(lenses)
        rng = f"B2:{last_col}{last_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FFF4F4F5",
                mid_type="percentile", mid_value=50, mid_color="FF6EE7B7",
                end_type="max", end_color="FF047857",
            ),
        )

    # Tab 8: Sources index
    ws = wb.create_sheet("Sources")
    headers = ["Citations", "Host", "URL"]
    _setup_sheet(ws, headers, [12, 28, 80])
    for r_idx, s in enumerate(snapshot.source_index, start=2):
        ws.cell(row=r_idx, column=1, value=s.get("citations", 0)).font = BODY_FONT
        ws.cell(row=r_idx, column=2, value=s.get("host", "")).font = BODY_FONT
        _link_cell(ws, r_idx, 3, s.get("url", ""))

    # Tab 9: Methodology — agent sessions
    ws = wb.create_sheet("Methodology")
    headers = ["Agent", "Started", "Completed", "Items completed", "Items failed", "Knowledge added"]
    _setup_sheet(ws, headers, [22, 18, 18, 16, 14, 16])
    for r_idx, s in enumerate(snapshot.agent_sessions, start=2):
        ws.cell(row=r_idx, column=1, value=s.get("agent_type", "")).font = BODY_FONT
        ws.cell(row=r_idx, column=2, value=(s.get("started_at") or "")[:16]).font = BODY_FONT
        ws.cell(row=r_idx, column=3, value=(s.get("completed_at") or "")[:16]).font = BODY_FONT
        ws.cell(row=r_idx, column=4, value=s.get("items_completed", 0)).font = BODY_FONT
        ws.cell(row=r_idx, column=5, value=s.get("items_failed", 0)).font = BODY_FONT
        ws.cell(row=r_idx, column=6, value=s.get("knowledge_added", 0)).font = BODY_FONT

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
